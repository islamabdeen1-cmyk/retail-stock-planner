import pandas as pd
import re
import os
import sys
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, messagebox
from ttkbootstrap import Style
import ttkbootstrap as ttk
from PIL import Image, ImageTk
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def resource_path(relative_path):
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.abspath("."), relative_path)

def clean_product_name(name):
    return re.sub(r'\s+\d+(\.\d+)?\s*LE$', '', str(name).strip())

def process_file(file_path, start_date, end_date, future_days):
    try:
        df = pd.read_excel(file_path)
        df['Clean Product'] = df['Product'].apply(clean_product_name)

        grouped = df.groupby(['Clean Product', 'Type', 'Warehouse']).agg({
            'Sold Quantity': 'sum',
            'Current Stock': 'sum',
            'Price': lambda x: (
                str(int(x.iloc[0])) if len(set(x.dropna())) == 1 and x.iloc[0] != 0
                else ', '.join(
                    str(int(v)) if float(v).is_integer() else str(v)
                    for v in sorted(set(x.dropna())) if float(v) != 0
                )
            )
        }).reset_index()

        days = (end_date - start_date).days or 1
        grouped['Daily Usage'] = grouped['Sold Quantity'] / days
        grouped['Needed for next days'] = grouped['Daily Usage'] * future_days
        grouped['Stock Difference'] = grouped['Current Stock'] - grouped['Needed for next days']

        grouped = grouped.round(2)

        # ترتيب الأعمدة زي ما طلبت
        columns_order = [
            'Warehouse', 'Clean Product', 'Price', 'Current Stock',
            'Sold Quantity', 'Daily Usage', 'Needed for next days',
            'Stock Difference', 'Type'
        ]
        grouped = grouped[columns_order]

        # اسم مميز للملف الناتج
        base_name = "نتيجة الكميات المطلوبة"
        extension = ".xlsx"
        counter = 0
        while True:
            filename = f"{base_name}{f' ({counter})' if counter else ''}{extension}"
            output_file = os.path.join(os.path.dirname(file_path), filename)
            if not os.path.exists(output_file):
                break
            counter += 1

        # حفظ الملف
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            grouped.to_excel(writer, index=False, sheet_name="نتيجة الكميات المطلوبة")

        # تنسيق الأعمدة والصفوف
        wb = load_workbook(output_file)
        ws = wb["نتيجة الكميات المطلوبة"]
        for column in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column)
            col_letter = get_column_letter(column[0].column)
            ws.column_dimensions[col_letter].width = max_length + 5
        for row in ws.iter_rows(min_row=2):
            ws.row_dimensions[row[0].row].height = 25
        wb.save(output_file)

        messagebox.showinfo("تم", f"تم إنشاء الملف:\n{output_file}")
    except Exception as e:
        messagebox.showerror("خطأ", f"حصل خطأ أثناء المعالجة:\n{e}")

def browse_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
    if file_path:
        file_entry.delete(0, tk.END)
        file_entry.insert(0, file_path)

def run_processing():
    file_path = file_entry.get()
    try:
        start = datetime.strptime(start_date_entry.get(), "%d-%m-%Y")
        end = datetime.strptime(end_date_entry.get(), "%d-%m-%Y")
        future_days = int(days_entry.get())
    except:
        messagebox.showerror("خطأ في البيانات", "اكتب التواريخ كده: يوم-شهر-سنة (مثال: 18-06-2025) وعدد الأيام يكون رقم.")
        return

    process_file(file_path, start, end, future_days)

# واجهة البرنامج
style = Style("flatly")
root = style.master
root.title("Noha Stock Planner")
root.geometry("520x520")

try:
    image = Image.open(resource_path("logo.png"))
    image = image.resize((120, 120))
    photo = ImageTk.PhotoImage(image)
    img_label = tk.Label(root, image=photo)
    img_label.image = photo
    img_label.pack(pady=5)
except:
    pass

ttk.Label(root, text="Noha Pharmacy", font=("Arial", 18, "bold")).pack()
ttk.Label(root, text="By Islam Abdeen", font=("Arial", 10)).pack(pady=5)

ttk.Label(root, text="اختيار ملف الإكسيل:").pack()
file_entry = ttk.Entry(root, width=60)
file_entry.pack(pady=2)
ttk.Button(root, text="📁 اختيار ملف", command=browse_file).pack(pady=5)

ttk.Label(root, text="تاريخ البداية (يوم-شهر-سنة):").pack()
start_date_entry = ttk.Entry(root)
start_date_entry.pack(pady=2)

ttk.Label(root, text="تاريخ النهاية (يوم-شهر-سنة):").pack()
end_date_entry = ttk.Entry(root)
end_date_entry.pack(pady=2)

ttk.Label(root, text="عدد أيام التغطية المطلوبة:").pack()
days_entry = ttk.Entry(root)
days_entry.pack(pady=2)

ttk.Button(root, text="🔍 احسب الكميات المطلوبة", bootstyle="success", command=run_processing).pack(pady=15)

root.mainloop()
